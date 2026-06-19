"""
Export full financial analysis workbook to Excel for learning.
Raw data (02–05) contains values; analysis sheets (06–20) use Excel formulas
that reference raw tabs and the assumptions sheet.
"""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from scripts.excel_formula_builder import (
    add_formula_analysis_sheets,
    build_refs,
    write_assumptions_sheet,
)
from src.budgeting import BudgetAnalyzer
from src.capital_structure import CapitalStructureAnalyzer
from src.cash_flow_analysis import CashFlowAnalyzer
from src.data_fetcher import FinancialDataFetcher
from src.dupont_analysis import DuPontAnalyzer
from src.executive_summary import ExecutiveSummaryGenerator
from src.forecasting import FinancialForecaster
from src.margin_bridge import MarginBridgeAnalyzer
from src.pl_analysis import PLAnalyzer
from src.sensitivity import SensitivityAnalyzer
from src.working_capital import WorkingCapitalAnalyzer

DESKTOP_PATHS = [
    Path.home() / "Desktop",
    Path(r"f:\桌面"),
]

SHEET_ORDER = [
    "00_Guide",
    "00_Pipeline",
    "01_Company_Info",
    "02_Income_Raw",
    "03_Balance_Raw",
    "04_CashFlow_Raw",
    "05_Price_1Y",
    "15a_Assumptions",
    "06_PL_Key_Lines",
    "07_Margins",
    "08_YoY_Growth",
    "09_Expense_Pct",
    "10_Margin_Bridge",
    "11_CashFlow_FCF",
    "12_Working_Capital",
    "13_DuPont",
    "14_Capital_Structure",
    "15_Budget",
    "16_Variance",
    "16b_Budget_Summary",
    "17_Forecast",
    "18_Scenarios",
    "19_Sensitivity_1Way",
    "19b_Sensitivity_2Way",
    "20_Tornado",
    "21_Executive_Summary",
]

GUIDE_ROWS = [
    ["Tesla Financial Analysis — Excel Learning Workbook", ""],
    ["", ""],
    ["如何使用本文件", ""],
    ["1. 02–05 页为 Yahoo Finance 原始数据（数值，勿改结构）", ""],
    ["   格式：行 = 科目，列 = 年份", ""],
    ["2. 15a_Assumptions 可修改增长率等假设，分析表会自动重算", ""],
    ["3. 06–20 页均为 Excel 公式，引用 02–04 原始表 + 假设页", ""],
    ["   点击单元格查看公式，便于学习 FP&A 建模", ""],
    ["4. 21_Executive_Summary 为 Python 自动生成的文字结论", ""],
    ["", ""],
    ["分析流程 (Pipeline)", "对应模块"],
    ["Step 1: 数据获取", "src/data_fetcher.py → 02–05"],
    ["Step 2: P&L 分析", "src/pl_analysis.py → 06–09"],
    ["Step 3: 净利润 Bridge", "src/margin_bridge.py → 10"],
    ["Step 4: 现金流 & FCF", "src/cash_flow_analysis.py → 11"],
    ["Step 5: 营运资本 & CCC", "src/working_capital.py → 12"],
    ["Step 6: DuPont 回报率", "src/dupont_analysis.py → 13"],
    ["Step 7: 资本结构", "src/capital_structure.py → 14"],
    ["Step 8: 预算编制", "src/budgeting.py → 15, 15a"],
    ["Step 9: 预算差异", "src/budgeting.py → 16"],
    ["Step 10: 收入预测", "src/forecasting.py → 17–18"],
    ["Step 11: 敏感性分析", "src/sensitivity.py → 19–20"],
    ["Step 12: 管理层摘要", "src/executive_summary.py → 21"],
    ["", ""],
    ["关键 FP&A 公式速查", ""],
    ["毛利率", "= 毛利 / 营收"],
    ["净利率", "= 净利润 / 营收"],
    ["FCF", "= 经营现金流 - CapEx"],
    ["ROE (DuPont)", "= 净利率 × 资产周转率 × 权益乘数"],
    ["CCC", "= DSO + DIO - DPO"],
    ["预算差异 %", "= (实际 - 预算) / 预算"],
]

PIPELINE_ROWS = [
    ["Step", "Module", "Input Data", "Output Metrics", "FP&A Purpose"],
    [1, "data_fetcher", "Yahoo Finance API", "Income / BS / CF / Prices", "获取标准化财报数据"],
    [2, "pl_analysis", "Income Statement", "Margins, YoY Growth, Expense %", "盈利能力与费用结构"],
    [3, "margin_bridge", "Income Statement (2 years)", "Revenue/Margin/OpEx effects", "解释净利润变动原因"],
    [4, "cash_flow_analysis", "Cash Flow + Income", "OCF, FCF, OCF/NI", "盈利质量与流动性"],
    [5, "working_capital", "BS + Income", "DSO, DIO, DPO, CCC", "营运资本效率"],
    [6, "dupont_analysis", "BS + Income", "ROE, ROA, ROIC decomposition", "股东回报驱动因素"],
    [7, "capital_structure", "BS + Income", "Net Debt, Leverage, Coverage", "资本结构与偿债能力"],
    [8, "budgeting", "Historical Income + Assumptions", "Forward budget", "年度预算编制"],
    [9, "budgeting (variance)", "Budget vs Actual", "Variance %, Favorable/Unfavorable", "预算执行监控"],
    [10, "forecasting", "Historical Revenue + Assumptions", "Base forecast + Scenarios", "滚动预测与情景规划"],
    [11, "sensitivity", "Latest year actuals", "Tornado, One-way tables", "风险与驱动因素测试"],
    [12, "executive_summary", "All above", "Narrative report", "管理层汇报材料"],
]

INCOME_LINE_ORDER = [
    "Total Revenue", "Operating Revenue", "Cost Of Revenue", "Gross Profit",
    "Operating Expense", "Research And Development", "Selling General And Administration",
    "Operating Income", "EBITDA", "EBIT", "Interest Expense", "Interest Income",
    "Pretax Income", "Tax Provision", "Net Income", "Basic EPS", "Diluted EPS",
]


def _is_year_index(df: pd.DataFrame) -> bool:
    if df.empty:
        return False
    if df.index.name == "Year":
        return True
    try:
        return all(int(i) > 1900 for i in df.index)
    except (TypeError, ValueError):
        return False


def _vertical_format(df: pd.DataFrame, index_label: str = "Line Item") -> pd.DataFrame:
    out = df.T.copy()
    out.index.name = index_label
    try:
        out = out.reindex(columns=sorted(out.columns, key=lambda x: int(x)))
    except (TypeError, ValueError):
        pass
    return out


def _order_line_items(df: pd.DataFrame, priority: list[str]) -> pd.DataFrame:
    idx = list(df.index)
    ordered = [k for k in priority if k in idx]
    ordered += sorted(k for k in idx if k not in ordered)
    return df.reindex(ordered)


def _write_raw_df(
    writer: pd.ExcelWriter,
    sheet: str,
    df: pd.DataFrame,
    index_label: str = "Line Item",
    line_order: list[str] | None = None,
) -> None:
    name = sheet[:31]
    out = df.copy()
    if _is_year_index(out):
        out = _vertical_format(out, index_label=index_label)
        if line_order:
            out = _order_line_items(out, line_order)
        out.to_excel(writer, sheet_name=name, index=True)
    else:
        out.to_excel(writer, sheet_name=name, index=True)


def _reorder_sheets(wb, order: list[str]) -> None:
    for i, name in enumerate(order):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))


def _write_executive_summary(wb, sections: list[dict]) -> None:
    ws = wb.create_sheet("21_Executive_Summary")
    ws.cell(1, 1, "Section")
    ws.cell(1, 2, "Narrative")
    for i, section in enumerate(sections, start=2):
        ws.cell(i, 1, section["title"])
        ws.cell(i, 2, section["body"])
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 100


def export_workbook(ticker: str = "TSLA", output_path: Path | None = None) -> Path:
    fetcher = FinancialDataFetcher(ticker)
    info = fetcher.get_company_info()
    income_stmt = fetcher.get_income_statement()
    balance_sheet = fetcher.get_balance_sheet()
    cash_flow = fetcher.get_cash_flow()
    price_history = fetcher.get_price_history()

    pl = PLAnalyzer(income_stmt)
    margins = pl.margin_analysis()
    growth = pl.yoy_growth()

    budget_analyzer = BudgetAnalyzer(income_stmt)
    base_year = int(income_stmt.index[-2])
    latest_year = int(income_stmt.index[-1])
    budget = budget_analyzer.create_budget(
        base_year=base_year,
        growth_assumptions={"Total Revenue": 0.12, "Operating Expense": 0.08},
        years_forward=2,
    )
    variance = budget_analyzer.variance_analysis(budget, latest_year) if latest_year in budget.index else None

    forecaster = FinancialForecaster(income_stmt)
    revenue_fc = forecaster.forecast_metric("Total Revenue", periods=3)
    scenarios = forecaster.scenario_forecast("Total Revenue", periods=3)

    sensitivity = SensitivityAnalyzer(income_stmt)
    tornado = sensitivity.tornado_data()

    cf_analyzer = CashFlowAnalyzer(income_stmt, cash_flow, balance_sheet)
    fcf_summary = cf_analyzer.fcf_summary()
    wc_summary = WorkingCapitalAnalyzer(income_stmt, balance_sheet).summary()
    dupont = DuPontAnalyzer(income_stmt, balance_sheet).decompose()
    cap_struct = CapitalStructureAnalyzer(income_stmt, balance_sheet).summary()
    bridge = MarginBridgeAnalyzer(income_stmt).yoy_bridge()

    exec_summary = ExecutiveSummaryGenerator(
        ticker=ticker,
        company_name=info["name"],
        income_stmt=income_stmt,
        margins=margins,
        growth=growth,
        fcf_summary=fcf_summary,
        wc_summary=wc_summary,
        dupont=dupont,
        cap_struct=cap_struct,
        variance=variance,
        bridge=bridge,
        tornado=tornado,
        revenue_fc=revenue_fc,
        scenarios=scenarios,
    ).generate()

    if output_path is None:
        stamp = datetime.now().strftime("%Y%m%d")
        for desktop in DESKTOP_PATHS:
            if desktop.exists():
                output_path = desktop / f"TSLA_Financial_Analysis_Workbook_{stamp}.xlsx"
                break
        else:
            output_path = ROOT / "outputs" / f"TSLA_Financial_Analysis_Workbook_{stamp}.xlsx"

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    company_df = pd.DataFrame([info]).T
    company_df.columns = ["Value"]

    price_export = price_history.tail(252).reset_index()
    date_col = price_export.columns[0]
    price_export[date_col] = pd.to_datetime(price_export[date_col]).dt.tz_localize(None)
    price_export.rename(columns={date_col: "Date"}, inplace=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        pd.DataFrame(GUIDE_ROWS, columns=["Topic", "Detail"]).to_excel(
            writer, sheet_name="00_Guide", index=False
        )
        pd.DataFrame(PIPELINE_ROWS[1:], columns=PIPELINE_ROWS[0]).to_excel(
            writer, sheet_name="00_Pipeline", index=False
        )
        company_df.to_excel(writer, sheet_name="01_Company_Info", index=True)
        _write_raw_df(writer, "02_Income_Raw", income_stmt, line_order=INCOME_LINE_ORDER)
        _write_raw_df(writer, "03_Balance_Raw", balance_sheet)
        _write_raw_df(writer, "04_CashFlow_Raw", cash_flow)
        price_export.to_excel(writer, sheet_name="05_Price_1Y", index=False)

    wb = load_workbook(output_path)
    refs = build_refs(wb)
    write_assumptions_sheet(wb, base_year, latest_year)
    add_formula_analysis_sheets(wb, refs, base_year, latest_year)
    _write_executive_summary(wb, exec_summary["sections"])
    _reorder_sheets(wb, SHEET_ORDER)
    wb.save(output_path)

    print("Workbook saved:", output_path.name)
    import shutil

    targets = [ROOT / output_path.name]
    for desktop in DESKTOP_PATHS:
        alt = desktop / output_path.name
        if desktop.exists() and alt.resolve() != output_path.resolve():
            targets.append(alt)

    for target in targets:
        target.parent.mkdir(parents=True, exist_ok=True)
        if target.resolve() != output_path.resolve():
            shutil.copy(output_path, target)
            print("Also copied to:", str(target))
    return output_path


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Output .xlsx path (default: desktop with today's date)",
    )
    args = parser.parse_args()
    path = export_workbook(output_path=args.output)
    print(str(path))
