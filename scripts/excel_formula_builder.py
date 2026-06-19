"""Build FP&A analysis sheets with Excel formulas referencing raw data tabs."""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Callable

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

INCOME_SHEET = "02_Income_Raw"
BALANCE_SHEET = "03_Balance_Raw"
CASHFLOW_SHEET = "04_CashFlow_Raw"
ASSUMPTIONS_SHEET = "15a_Assumptions"


@dataclass
class SheetRefs:
    income_rows: dict[str, int] = field(default_factory=dict)
    balance_rows: dict[str, int] = field(default_factory=dict)
    cashflow_rows: dict[str, int] = field(default_factory=dict)
    income_years: dict[int, int] = field(default_factory=dict)
    balance_years: dict[int, int] = field(default_factory=dict)
    cashflow_years: dict[int, int] = field(default_factory=dict)
    years: list[int] = field(default_factory=list)


def _ref(sheet: str, row: int, col: int, abs_row: bool = True, abs_col: bool = False) -> str:
    letter = get_column_letter(col)
    r = f"${row}" if abs_row else str(row)
    c = f"${letter}" if abs_col else letter
    return f"'{sheet}'!{c}{r}"


def build_refs(wb: Workbook) -> SheetRefs:
    refs = SheetRefs()
    refs.income_rows = _row_map(wb[INCOME_SHEET])
    refs.balance_rows = _row_map(wb[BALANCE_SHEET])
    refs.cashflow_rows = _row_map(wb[CASHFLOW_SHEET])
    refs.income_years = _year_cols(wb[INCOME_SHEET])
    refs.balance_years = _year_cols(wb[BALANCE_SHEET])
    refs.cashflow_years = _year_cols(wb[CASHFLOW_SHEET])
    refs.years = sorted(refs.income_years)
    return refs


def _row_map(ws: Worksheet, label_col: int = 1, start_row: int = 2) -> dict[str, int]:
    out: dict[str, int] = {}
    for row in range(start_row, ws.max_row + 1):
        val = ws.cell(row=row, column=label_col).value
        if val is not None and str(val).strip():
            out[str(val)] = row
    return out


def _year_cols(ws: Worksheet, header_row: int = 1, start_col: int = 2) -> dict[int, int]:
    out: dict[int, int] = {}
    for col in range(start_col, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is not None:
            try:
                out[int(val)] = col
            except (TypeError, ValueError):
                pass
    return out


class FormulaRefs:
    def __init__(self, refs: SheetRefs):
        self.r = refs

    def _row(self, mapping: dict[str, int], item: str) -> int | None:
        return mapping.get(item)

    def inc(self, item: str, year: int, abs_col: bool = False) -> str:
        row = self._row(self.r.income_rows, item)
        col = self.r.income_years.get(year)
        if row is None or col is None:
            return '""'
        return _ref(INCOME_SHEET, row, col, abs_col=abs_col)

    def bs(self, item: str, year: int, abs_col: bool = False) -> str:
        row = self._row(self.r.balance_rows, item)
        col = self.r.balance_years.get(year)
        if row is None or col is None:
            return '""'
        return _ref(BALANCE_SHEET, row, col, abs_col=abs_col)

    def cf(self, item: str, year: int, abs_col: bool = False) -> str:
        row = self._row(self.r.cashflow_rows, item)
        col = self.r.cashflow_years.get(year)
        if row is None or col is None:
            return '""'
        return _ref(CASHFLOW_SHEET, row, col, abs_col=abs_col)

    def prev_year(self, year: int) -> int | None:
        years = self.r.years
        if year not in years:
            return None
        idx = years.index(year)
        return years[idx - 1] if idx > 0 else None

    def has_inc(self, item: str) -> bool:
        return item in self.r.income_rows

    def has_bs(self, item: str) -> bool:
        return item in self.r.balance_rows


def _write_header(ws: Worksheet, years: list[int], label: str = "Metric") -> None:
    ws.cell(1, 1, label)
    for i, year in enumerate(years):
        ws.cell(1, i + 2, year)


def _write_metric_rows(
    ws: Worksheet,
    metrics: list[str],
    years: list[int],
    formula_fn: Callable[[str, int], str],
    label: str = "Metric",
) -> None:
    _write_header(ws, years, label)
    for i, metric in enumerate(metrics):
        row = i + 2
        ws.cell(row, 1, metric)
        for j, year in enumerate(years):
            ws.cell(row, j + 2, formula_fn(metric, year))


def write_assumptions_sheet(wb: Workbook, base_year: int, latest_year: int) -> None:
    ws = wb.create_sheet(ASSUMPTIONS_SHEET)
    ws.cell(1, 1, "Parameter")
    ws.cell(1, 2, "Value")
    ws.cell(1, 3, "Notes")

    rows = [
        ("Base Year (budget)", base_year, "Last full-year actual used for forward budget"),
        ("Latest Actual Year", latest_year, "Year used for variance analysis"),
        ("Budget — Revenue Growth %", 0.12, "Applied to Total Revenue each forward year"),
        ("Budget — COGS Growth %", 0.12, "Applied to Cost Of Revenue"),
        ("Budget — OpEx Growth %", 0.08, "Applied to Operating Expense"),
        ("Budget — Net Income Tax Rate %", 0.25, "Net Income = Operating Income × (1 − rate)"),
        ("Budget — Years Forward", 2, "Number of budget years after base year"),
        ("Forecast — Revenue Growth %", 0.10, "Simple forward revenue forecast"),
        ("Scenario — Bear Growth %", -0.05, "Annual revenue growth in bear case"),
        ("Scenario — Base Growth %", 0.10, "Annual revenue growth in base case"),
        ("Scenario — Bull Growth %", 0.25, "Annual revenue growth in bull case"),
        ("Sensitivity — Driver Change %", 0.10, "± change used in tornado / one-way"),
    ]
    for i, (param, value, note) in enumerate(rows, start=2):
        ws.cell(i, 1, param)
        ws.cell(i, 2, value)
        ws.cell(i, 3, note)

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["C"].width = 48


def _assumption_cell(row: int) -> str:
    return f"'{ASSUMPTIONS_SHEET}'!$B${row}"


ASSUMP = {
    "base_year": 2,
    "latest_year": 3,
    "rev_growth": 4,
    "cogs_growth": 5,
    "opex_growth": 6,
    "tax_rate": 7,
    "years_forward": 8,
    "fc_growth": 9,
    "bear": 10,
    "base_scenario": 11,
    "bull": 12,
    "sens_pct": 13,
}


def add_formula_analysis_sheets(wb: Workbook, refs: SheetRefs, base_year: int, latest_year: int) -> None:
    f = FormulaRefs(refs)
    years = refs.years

    _sheet_pl_key_lines(wb, f, years)
    _sheet_margins(wb, f, years)
    _sheet_yoy_growth(wb, f, years)
    _sheet_expense_pct(wb, f, years)
    _sheet_margin_bridge(wb, f, years)
    _sheet_cashflow_fcf(wb, f, years)
    _sheet_working_capital(wb, f, years)
    _sheet_dupont(wb, f, years)
    _sheet_capital_structure(wb, f, years)
    _sheet_budget(wb, f, base_year)
    _sheet_variance(wb, f, latest_year, base_year)
    _sheet_budget_summary(wb, f, years, base_year, latest_year)
    _sheet_forecast(wb, f, years, latest_year)
    _sheet_scenarios(wb, f, latest_year)
    _sheet_sensitivity_one_way(wb, f, latest_year)
    _sheet_sensitivity_two_way(wb, f, latest_year)
    _sheet_tornado(wb, f, latest_year)


def _sheet_pl_key_lines(wb: Workbook, f: FormulaRefs, years: list[int]) -> None:
    ws = wb.create_sheet("06_PL_Key_Lines")
    items = [
        "Total Revenue", "Cost Of Revenue", "Gross Profit",
        "Operating Expense", "Operating Income", "EBITDA", "Net Income",
    ]
    items = [i for i in items if f.has_inc(i)]

    def formula(item: str, year: int) -> str:
        return f"=IFERROR({f.inc(item, year)},"")"

    _write_metric_rows(ws, items, years, formula, label="Line Item")


def _sheet_margins(wb: Workbook, f: FormulaRefs, years: list[int]) -> None:
    ws = wb.create_sheet("07_Margins")
    metrics = ["Gross Margin %", "Operating Margin %", "Net Margin %", "EBITDA Margin %"]
    numerators = {
        "Gross Margin %": "Gross Profit",
        "Operating Margin %": "Operating Income",
        "Net Margin %": "Net Income",
        "EBITDA Margin %": "EBITDA",
    }

    def formula(metric: str, year: int) -> str:
        num = numerators[metric]
        if metric == "EBITDA Margin %" and not f.has_inc("EBITDA"):
            return '=""'
        return (
            f'=IFERROR(IF({f.inc("Total Revenue", year)}=0,"",'
            f'{f.inc(numerators[metric], year)}/{f.inc("Total Revenue", year)}*100),"")'
        )

    _write_metric_rows(ws, metrics, years, formula)


def _sheet_yoy_growth(wb: Workbook, f: FormulaRefs, years: list[int]) -> None:
    ws = wb.create_sheet("08_YoY_Growth")
    items = [
        "Total Revenue", "Cost Of Revenue", "Gross Profit",
        "Operating Expense", "Operating Income", "Net Income", "EBITDA",
    ]
    items = [i for i in items if f.has_inc(i)]
    display = [f"{i} YoY %" for i in items]
    item_map = dict(zip(display, items))

    def formula(label: str, year: int) -> str:
        item = item_map[label]
        prev = f.prev_year(year)
        if prev is None:
            return '=""'
        return (
            f'=IFERROR(IF({f.inc(item, prev)}=0,"",'
            f'({f.inc(item, year)}/{f.inc(item, prev)}-1)*100),"")'
        )

    _write_metric_rows(ws, display, years, formula)


def _sheet_expense_pct(wb: Workbook, f: FormulaRefs, years: list[int]) -> None:
    ws = wb.create_sheet("09_Expense_Pct")
    candidates = [
        ("COGS % of Revenue", "Cost Of Revenue"),
        ("OpEx % of Revenue", "Operating Expense"),
        ("R&D % of Revenue", "Research And Development"),
        ("SG&A % of Revenue", "Selling General And Administration"),
    ]
    metrics = [label for label, item in candidates if f.has_inc(item)]
    item_map = {label: item for label, item in candidates if f.has_inc(item)}

    def formula(metric: str, year: int) -> str:
        item = item_map[metric]
        return (
            f'=IFERROR(IF({f.inc("Total Revenue", year)}=0,"",'
            f'{f.inc(item, year)}/{f.inc("Total Revenue", year)}*100),"")'
        )

    _write_metric_rows(ws, metrics, years, formula)


def _sheet_margin_bridge(wb: Workbook, f: FormulaRefs, years: list[int]) -> None:
    ws = wb.create_sheet("10_Margin_Bridge")
    bridge_years = [y for y in years if f.prev_year(y) is not None]
    metrics = ["Revenue Volume", "Margin / Mix", "OpEx Change", "Below-the-Line", "Total Δ Net Income"]
    _write_header(ws, bridge_years, label="Line Item")

    for i, metric in enumerate(metrics):
        row = i + 2
        ws.cell(row, 1, metric)
        for j, year in enumerate(bridge_years):
            prev = f.prev_year(year)
            rev_c, rev_p = f.inc("Total Revenue", year), f.inc("Total Revenue", prev)
            gp_c, gp_p = f.inc("Gross Profit", year), f.inc("Gross Profit", prev)
            oi_c, oi_p = f.inc("Operating Income", year), f.inc("Operating Income", prev)
            ni_c, ni_p = f.inc("Net Income", year), f.inc("Net Income", prev)

            if metric == "Revenue Volume":
                formula = f'=IFERROR(({rev_c}-{rev_p})*IF({rev_p}=0,0,{gp_p}/{rev_p}),0)'
            elif metric == "Margin / Mix":
                formula = f'=IFERROR(IF(OR({rev_c}=0,{rev_p}=0),0,{rev_c}*({gp_c}/{rev_c}-{gp_p}/{rev_p})),0)'
            elif metric == "OpEx Change":
                formula = f'=IFERROR(({oi_c}-{oi_p})-({gp_c}-{gp_p}),0)'
            elif metric == "Below-the-Line":
                formula = f'=IFERROR(({ni_c}-{ni_p})-({oi_c}-{oi_p}),0)'
            else:
                formula = f'=IFERROR({ni_c}-{ni_p},0)'
            ws.cell(row, j + 2, formula)


def _sheet_cashflow_fcf(wb: Workbook, f: FormulaRefs, years: list[int]) -> None:
    ws = wb.create_sheet("11_CashFlow_FCF")
    cf_years = sorted(set(years) & set(f.r.cashflow_years))
    metrics = [
        "Operating Cash Flow", "Capital Expenditure", "Free Cash Flow",
        "FCF Margin %", "CapEx % of Revenue", "OCF / Net Income",
    ]

    def formula(metric: str, year: int) -> str:
        ocf = f.cf("Operating Cash Flow", year)
        capex = f.cf("Capital Expenditure", year)
        rev = f.inc("Total Revenue", year)
        ni = f.inc("Net Income", year)
        if metric == "Operating Cash Flow":
            return f"=IFERROR({ocf},0)"
        if metric == "Capital Expenditure":
            return f"=IFERROR(ABS({capex}),0)"
        if metric == "Free Cash Flow":
            fcf = f.cf("Free Cash Flow", year)
            return f'=IFERROR(IF({fcf}=0,{ocf}-ABS({capex}),{fcf}),0)'
        if metric == "FCF Margin %":
            fcf_formula = f'IFERROR(IF({f.cf("Free Cash Flow", year)}=0,{ocf}-ABS({capex}),{f.cf("Free Cash Flow", year)}),0)'
            return f'=IFERROR(IF({rev}=0,"",({fcf_formula})/{rev}*100),"")'
        if metric == "CapEx % of Revenue":
            return f'=IFERROR(IF({rev}=0,"",ABS({capex})/{rev}*100),"")'
        return f'=IFERROR(IF({ni}=0,"",{ocf}/{ni}),"")'

    _write_metric_rows(ws, metrics, cf_years, formula)


def _sheet_working_capital(wb: Workbook, f: FormulaRefs, years: list[int]) -> None:
    ws = wb.create_sheet("12_Working_Capital")
    bs_years = sorted(set(years) & set(f.r.balance_years))
    metrics = [
        "Working Capital", "WC % of Revenue", "Δ Working Capital",
        "DSO (days)", "DIO (days)", "DPO (days)", "Cash Conversion Cycle",
    ]

    def formula(metric: str, year: int) -> str:
        wc = f.bs("Working Capital", year)
        rev = f.inc("Total Revenue", year)
        cogs = f.inc("Cost Of Revenue", year)
        ar = f.bs("Accounts Receivable", year)
        inv = f.bs("Inventory", year)
        ap = f.bs("Accounts Payable", year)
        prev = f.prev_year(year)
        prev_wc = f.bs("Working Capital", prev) if prev else "0"

        if metric == "Working Capital":
            return f"=IFERROR({wc},0)"
        if metric == "WC % of Revenue":
            return f'=IFERROR(IF({rev}=0,"",{wc}/{rev}*100),"")'
        if metric == "Δ Working Capital":
            if prev is None:
                return '=""'
            return f"=IFERROR({wc}-{prev_wc},\"\")"
        if metric == "DSO (days)":
            return f'=IFERROR(IF({rev}=0,"",{ar}/{rev}*365),"")'
        if metric == "DIO (days)":
            return f'=IFERROR(IF({cogs}=0,"",{inv}/{cogs}*365),"")'
        if metric == "DPO (days)":
            return f'=IFERROR(IF({cogs}=0,"",{ap}/{cogs}*365),"")'
        dso = f'IFERROR(IF({rev}=0,0,{ar}/{rev}*365),0)'
        dio = f'IFERROR(IF({cogs}=0,0,{inv}/{cogs}*365),0)'
        dpo = f'IFERROR(IF({cogs}=0,0,{ap}/{cogs}*365),0)'
        return f"=IFERROR({dso}+{dio}-{dpo},\"\")"

    _write_metric_rows(ws, metrics, bs_years, formula)


def _sheet_dupont(wb: Workbook, f: FormulaRefs, years: list[int]) -> None:
    ws = wb.create_sheet("13_DuPont")
    bs_years = sorted(set(years) & set(f.r.balance_years))
    metrics = [
        "ROE %", "ROA %", "ROIC %",
        "Net Margin %", "Asset Turnover", "Equity Multiplier",
    ]

    def equity_ref(year: int) -> str:
        if f.has_bs("Stockholders Equity"):
            return f.bs("Stockholders Equity", year)
        return f.bs("Common Stock Equity", year)

    def formula(metric: str, year: int) -> str:
        rev = f.inc("Total Revenue", year)
        ni = f.inc("Net Income", year)
        ebit = f.inc("Operating Income", year)
        assets = f.bs("Total Assets", year)
        equity = equity_ref(year)
        invested = f.bs("Invested Capital", year)
        nm = f"IFERROR({ni}/{rev},0)"
        at = f"IFERROR({rev}/{assets},0)"
        em = f"IFERROR({assets}/{equity},0)"
        if metric == "Net Margin %":
            return f'=IFERROR(IF({rev}=0,"",{ni}/{rev}*100),"")'
        if metric == "Asset Turnover":
            return f"=IFERROR({at},\"\")"
        if metric == "Equity Multiplier":
            return f"=IFERROR({em},\"\")"
        if metric == "ROE %":
            return f"=IFERROR(({nm})*({at})*({em})*100,\"\")"
        if metric == "ROA %":
            return f'=IFERROR(IF({assets}=0,"",{ni}/{assets}*100),"")'
        return f'=IFERROR(IF({invested}=0,"",{ebit}/{invested}*100),"")'

    _write_metric_rows(ws, metrics, bs_years, formula)


def _sheet_capital_structure(wb: Workbook, f: FormulaRefs, years: list[int]) -> None:
    ws = wb.create_sheet("14_Capital_Structure")
    bs_years = sorted(set(years) & set(f.r.balance_years))
    metrics = [
        "Total Debt", "Net Debt", "Cash",
        "Debt / Equity", "Net Debt / EBITDA", "Interest Coverage",
    ]

    def cash_ref(year: int) -> str:
        if f.has_bs("Cash And Cash Equivalents"):
            return f.bs("Cash And Cash Equivalents", year)
        return f.bs("Cash Cash Equivalents And Short Term Investments", year)

    def equity_ref(year: int) -> str:
        if f.has_bs("Stockholders Equity"):
            return f.bs("Stockholders Equity", year)
        return f.bs("Common Stock Equity", year)

    def formula(metric: str, year: int) -> str:
        debt = f.bs("Total Debt", year)
        cash = cash_ref(year)
        equity = equity_ref(year)
        ebitda = f.inc("EBITDA", year)
        ebit = f.inc("Operating Income", year)
        interest = f.inc("Interest Expense", year)
        net_debt = f"({debt})-({cash})"
        if metric == "Total Debt":
            return f"=IFERROR({debt},0)"
        if metric == "Cash":
            return f"=IFERROR({cash},0)"
        if metric == "Net Debt":
            return f"=IFERROR({net_debt},0)"
        if metric == "Debt / Equity":
            return f'=IFERROR(IF({equity}=0,"",{debt}/{equity}),"")'
        if metric == "Net Debt / EBITDA":
            return f'=IFERROR(IF({ebitda}=0,"",({net_debt})/{ebitda}),"")'
        return f'=IFERROR(IF(ABS({interest})=0,"",{ebit}/ABS({interest})),"")'

    _write_metric_rows(ws, metrics, bs_years, formula)


def _sheet_budget(wb: Workbook, f: FormulaRefs, base_year: int) -> None:
    ws = wb.create_sheet("15_Budget")
    ws.cell(1, 1, "Line Item")
    ws.cell(1, 2, "Formula / Notes")

    budget_years = [base_year + 1, base_year + 2]
    for i, year in enumerate(budget_years):
        ws.cell(1, i + 3, year)

    metrics = [
        "Total Revenue", "Cost Of Revenue", "Gross Profit",
        "Operating Expense", "Operating Income", "Net Income",
    ]
    metrics = [m for m in metrics if f.has_inc(m)]

    for i, metric in enumerate(metrics):
        row = i + 2
        ws.cell(row, 1, metric)
        ws.cell(row, 2, f"= prior year × (1 + growth)^n from {ASSUMPTIONS_SHEET}")

        for j, year in enumerate(budget_years):
            n = year - base_year
            prev_year = year - 1
            if prev_year == base_year:
                base_ref = f.inc(metric, base_year)
            else:
                base_ref = _ref("15_Budget", row, j + 2, abs_row=True, abs_col=False)

            if metric == "Gross Profit":
                rev_row = metrics.index("Total Revenue") + 2
                cogs_row = metrics.index("Cost Of Revenue") + 2
                col = j + 3
                formula = f"=IFERROR({_ref('15_Budget', rev_row, col)}-{_ref('15_Budget', cogs_row, col)},0)"
            elif metric == "Operating Income":
                gp_row = metrics.index("Gross Profit") + 2
                opex_row = metrics.index("Operating Expense") + 2
                col = j + 3
                formula = f"=IFERROR({_ref('15_Budget', gp_row, col)}-{_ref('15_Budget', opex_row, col)},0)"
            elif metric == "Net Income":
                oi_row = metrics.index("Operating Income") + 2
                col = j + 3
                tax = _assumption_cell(ASSUMP["tax_rate"])
                formula = f"=IFERROR({_ref('15_Budget', oi_row, col)}*(1-{tax}),0)"
            else:
                growth_row = {
                    "Total Revenue": ASSUMP["rev_growth"],
                    "Cost Of Revenue": ASSUMP["cogs_growth"],
                    "Operating Expense": ASSUMP["opex_growth"],
                }[metric]
                formula = (
                    f"=IFERROR({base_ref}*(1+{_assumption_cell(growth_row)})^{n},0)"
                )
            ws.cell(row, j + 3, formula)


def _sheet_variance(wb: Workbook, f: FormulaRefs, latest_year: int, base_year: int) -> None:
    ws = wb.create_sheet("16_Variance")
    metrics = [
        "Total Revenue", "Gross Profit", "Operating Income", "Net Income",
    ]
    metrics = [m for m in metrics if f.has_inc(m)]
    headers = ["Metric", "Budget", "Actual", "Variance", "Variance %", "Status"]
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i, h)

    budget_col = 3 + (latest_year - base_year - 1)
    for i, metric in enumerate(metrics):
        row = i + 2
        budget_row = metrics.index(metric) + 2
        bud = _ref("15_Budget", budget_row, budget_col)
        act = f.inc(metric, latest_year)
        ws.cell(row, 1, metric)
        ws.cell(row, 2, f"=IFERROR({bud},0)")
        ws.cell(row, 3, f"=IFERROR({act},0)")
        ws.cell(row, 4, f"=IFERROR(C{row}-B{row},0)")
        ws.cell(row, 5, f'=IFERROR(IF(B{row}=0,"",D{row}/B{row}*100),"")')
        ws.cell(row, 6, f'=IF(D{row}>=0,"Favorable","Unfavorable")')


def _sheet_budget_summary(wb: Workbook, f: FormulaRefs, years: list[int], base_year: int, latest_year: int) -> None:
    ws = wb.create_sheet("16b_Budget_Summary")
    hist_years = [y for y in years if y <= latest_year]
    budget_years = [base_year + 1, base_year + 2]
    all_years = hist_years + [y for y in budget_years if y not in hist_years]
    metrics = [
        "Total Revenue", "Cost Of Revenue", "Gross Profit",
        "Operating Income", "Net Income",
    ]
    metrics = [m for m in metrics if f.has_inc(m)]
    _write_header(ws, all_years, label="Line Item")

    for i, metric in enumerate(metrics):
        row = i + 2
        ws.cell(row, 1, metric)
        budget_row = metrics.index(metric) + 2
        for j, year in enumerate(all_years):
            if year in hist_years:
                ws.cell(row, j + 2, f"=IFERROR({f.inc(metric, year)},0)")
            else:
                bcol = 3 if year == budget_years[0] else 4
                ws.cell(row, j + 2, f"=IFERROR({_ref('15_Budget', budget_row, bcol)},0)")


def _sheet_forecast(wb: Workbook, f: FormulaRefs, years: list[int], latest_year: int) -> None:
    ws = wb.create_sheet("17_Forecast")
    forecast_years = [latest_year + 1, latest_year + 2, latest_year + 3]
    all_years = years + forecast_years
    _write_header(ws, all_years, label="Line Item")
    ws.cell(2, 1, "Total Revenue")
    growth = _assumption_cell(ASSUMP["fc_growth"])

    for j, year in enumerate(all_years):
        col = j + 2
        if year in years:
            ws.cell(2, col, f"=IFERROR({f.inc('Total Revenue', year)},0)")
        else:
            n = year - latest_year
            prev_col = get_column_letter(col - 1)
            ws.cell(2, col, f"=IFERROR({prev_col}2*(1+{growth})^{n},0)")


def _sheet_scenarios(wb: Workbook, f: FormulaRefs, latest_year: int) -> None:
    ws = wb.create_sheet("18_Scenarios")
    forecast_years = [latest_year + 1, latest_year + 2, latest_year + 3]
    _write_header(ws, forecast_years, label="Scenario")
    base_rev = f.inc("Total Revenue", latest_year)
    scenarios = [
        ("Bear", ASSUMP["bear"]),
        ("Base", ASSUMP["base_scenario"]),
        ("Bull", ASSUMP["bull"]),
    ]
    for i, (name, rate_row) in enumerate(scenarios):
        row = i + 2
        ws.cell(row, 1, name)
        rate = _assumption_cell(rate_row)
        for j, year in enumerate(forecast_years):
            n = year - latest_year
            ws.cell(row, j + 2, f"=IFERROR({base_rev}*(1+{rate})^{n},0)")


def _sheet_sensitivity_one_way(wb: Workbook, f: FormulaRefs, latest_year: int) -> None:
    ws = wb.create_sheet("19_Sensitivity_1Way")
    changes = [-20, -15, -10, -5, 0, 5, 10, 15, 20]
    ws.cell(1, 1, "Metric")
    for j, chg in enumerate(changes):
        ws.cell(1, j + 2, f"{chg}%")

    base_rev = f.inc("Total Revenue", latest_year)
    base_ni = f.inc("Net Income", latest_year)
    base_oi = f.inc("Operating Income", latest_year)
    base_gp = f.inc("Gross Profit", latest_year)
    base_cogs = f.inc("Cost Of Revenue", latest_year)

    rows = [
        ("Adjusted Net Income", "ni"),
        ("Net Income Impact", "impact"),
        ("Net Income Impact %", "impact_pct"),
    ]
    for i, (label, kind) in enumerate(rows):
        row = i + 2
        ws.cell(row, 1, label)
        for j, chg in enumerate(changes):
            col = j + 2
            pct = chg / 100
            adj_rev = f"({base_rev})*(1+{pct})"
            rev_ratio = f"IF(({base_rev})=0,1,{adj_rev}/({base_rev}))"
            adj_cogs = f"({base_cogs})*({rev_ratio})"
            adj_gp = f"({adj_rev})-({adj_cogs})"
            adj_oi = f"({adj_gp})-(({f.inc('Operating Expense', latest_year)}))"
            adj_ni = f"IF(({base_oi})=0,({adj_oi})*0.75,({adj_oi})*(({base_ni})/({base_oi})))"
            if kind == "ni":
                ws.cell(row, col, f"=IFERROR({adj_ni},0)")
            elif kind == "impact":
                ws.cell(row, col, f"=IFERROR({adj_ni}-({base_ni}),0)")
            else:
                ws.cell(row, col, f'=IFERROR(IF(({base_ni})=0,"",(({adj_ni})-({base_ni}))/({base_ni})*100),"")')


def _sheet_sensitivity_two_way(wb: Workbook, f: FormulaRefs, latest_year: int) -> None:
    ws = wb.create_sheet("19b_Sensitivity_2Way")
    changes = [-15, -10, -5, 0, 5, 10, 15]
    ws.cell(1, 1, "OpEx Change")
    for j, cx in enumerate(changes):
        ws.cell(1, j + 2, f"Revenue {cx:+}%")

    base_rev = f.inc("Total Revenue", latest_year)
    base_ni = f.inc("Net Income", latest_year)
    base_oi = f.inc("Operating Income", latest_year)
    base_gp = f.inc("Gross Profit", latest_year)
    base_cogs = f.inc("Cost Of Revenue", latest_year)
    base_opex = f.inc("Operating Expense", latest_year)

    for i, cy in enumerate(changes):
        row = i + 2
        ws.cell(row, 1, f"{cy:+}%")
        for j, cx in enumerate(changes):
            pct_x, pct_y = cx / 100, cy / 100
            adj_rev = f"({base_rev})*(1+{pct_x})"
            rev_ratio = f"IF(({base_rev})=0,1,{adj_rev}/({base_rev}))"
            adj_cogs = f"({base_cogs})*({rev_ratio})"
            adj_gp = f"({adj_rev})-({adj_cogs})"
            adj_opex = f"({base_opex})*(1+{pct_y})"
            adj_oi = f"({adj_gp})-({adj_opex})"
            adj_ni = f"IF(({base_oi})=0,({adj_oi})*0.75,({adj_oi})*(({base_ni})/({base_oi})))"
            ws.cell(row, j + 2, f"=IFERROR({adj_ni},0)")


def _sheet_tornado(wb: Workbook, f: FormulaRefs, latest_year: int) -> None:
    ws = wb.create_sheet("20_Tornado")
    headers = ["Driver", "Low Case", "High Case", "Base", "Range"]
    for i, h in enumerate(headers, start=1):
        ws.cell(1, i, h)

    sens = _assumption_cell(ASSUMP["sens_pct"])
    base_ni = f.inc("Net Income", latest_year)
    base_oi = f.inc("Operating Income", latest_year)
    base_rev = f.inc("Total Revenue", latest_year)
    base_gp = f.inc("Gross Profit", latest_year)
    base_cogs = f.inc("Cost Of Revenue", latest_year)
    base_opex = f.inc("Operating Expense", latest_year)

    drivers = [
        ("Total Revenue", "rev"),
        ("Cost Of Revenue", "cogs"),
        ("Operating Expense", "opex"),
    ]
    drivers = [(d, k) for d, k in drivers if f.has_inc(d)]

    for i, (driver, kind) in enumerate(drivers):
        row = i + 2
        ws.cell(row, 1, driver)
        ws.cell(row, 4, f"=IFERROR({base_ni},0)")

        if kind == "rev":
            low_rev = f"({base_rev})*(1-{sens})"
            high_rev = f"({base_rev})*(1+{sens})"
            low_ratio = f"IF(({base_rev})=0,1,{low_rev}/({base_rev}))"
            high_ratio = f"IF(({base_rev})=0,1,{high_rev}/({base_rev}))"
            low_cogs = f"({base_cogs})*({low_ratio})"
            high_cogs = f"({base_cogs})*({high_ratio})"
            low_oi = f"(({low_rev})-({low_cogs}))-({base_opex})"
            high_oi = f"(({high_rev})-({high_cogs}))-({base_opex})"
        elif kind == "cogs":
            low_cogs = f"({base_cogs})*(1+{sens})"
            high_cogs = f"({base_cogs})*(1-{sens})"
            low_oi = f"({base_rev})-({low_cogs})-({base_opex})"
            high_oi = f"({base_rev})-({high_cogs})-({base_opex})"
        else:
            low_opex = f"({base_opex})*(1+{sens})"
            high_opex = f"({base_opex})*(1-{sens})"
            low_oi = f"({base_gp})-({low_opex})"
            high_oi = f"({base_gp})-({high_opex})"

        low_ni = f"IF(({base_oi})=0,({low_oi})*0.75,({low_oi})*(({base_ni})/({base_oi})))"
        high_ni = f"IF(({base_oi})=0,({high_oi})*0.75,({high_oi})*(({base_ni})/({base_oi})))"
        ws.cell(row, 2, f"=IFERROR({low_ni},0)")
        ws.cell(row, 3, f"=IFERROR({high_ni},0)")
        ws.cell(row, 5, f"=IFERROR(ABS(C{row}-B{row}),0)")
